from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


class ExcelService:
    def build_workbook(self, sheet_name: str, rows: list[dict], columns: list[tuple[str, str]]) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name[:31]
        sheet.append([label for label, _ in columns])
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F2937")

        for row in rows:
            sheet.append([row.get(key, "") for _, key in columns])

        for index, _ in enumerate(columns, start=1):
            letter = get_column_letter(index)
            width = 12
            for cell in sheet[letter]:
                width = max(width, len(str(cell.value or "")) + 2)
            sheet.column_dimensions[letter].width = min(width, 42)

        stream = BytesIO()
        workbook.save(stream)
        return stream.getvalue()
